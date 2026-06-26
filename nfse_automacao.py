"""
nfse_automacao.py - Motor principal da Automacao NFSe (execucao local).

Fluxo:
  1. Preparar parametros (datas do mes anterior ou customizadas)
  2. Ler certificados .pfx locais (CNPJ, senha, thumbprint)
  3. Para cada cliente, baixar os DF-e direto da API oficial ADN (mTLS com o
     certificado) via adn_client — sem navegador, extensao ou captcha
  4. Filtrar por competencia (mes) e salvar em PASTA_SAIDA/<Cliente>/<Tipo>/
  5. (Opcional) Importar os XMLs no Dominio Web
  6. Enviar relatorio final por e-mail
"""

from __future__ import annotations

import html as _html
import logging
import re
import smtplib
import threading
import time
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date, datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Optional

from config_defaults import apply_defaults, ensure_config_file

try:
    ensure_config_file()
    import config
except ImportError:
    ensure_config_file()
    import config
apply_defaults(config)
from cert_reader import CertificadoInfo, indexar_certificados_por_cnpj, listar_certificados


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("nfse")


class ExecucaoCancelada(RuntimeError):
    """Sinaliza que o usuario cancelou a automacao em andamento."""


def _check_cancel(cancel_event: threading.Event | None, mensagem: str = "") -> None:
    if cancel_event and cancel_event.is_set():
        raise ExecucaoCancelada(mensagem or "Execucao cancelada pelo usuario.")


@dataclass
class Parametros:
    xlsx_path: str
    pasta_certs: str
    pasta_saida: str
    data_inicio: str
    data_fim: str
    cnpjs: Optional[list[str]] = field(default=None)
    # Tipos de nota a baixar: "emitidas", "recebidas" ou "ambas".
    tipos: str = "ambas"
    # Mapa CNPJ -> nome vindo da aba Clientes (servidor/por maquina). Quando
    # presente, e usado como filtro/origem de nomes NO LUGAR do XLSX local.
    clientes_map: Optional[dict[str, str]] = field(default=None)


@dataclass
class ResultadoCNPJ:
    cnpj: str
    nome: str
    status: str
    notas_emitidas: int = 0
    notas_recebidas: int = 0
    arquivo_emitidas: str = ""
    arquivo_recebidas: str = ""
    thumbprint: str = ""
    erro: str = ""
    importado_dominio: bool = False

    @property
    def notas_encontradas(self) -> int:
        return self.notas_emitidas + self.notas_recebidas


@dataclass
class Cliente:
    cnpj: str
    nome: str


class NFSeRunner:
    """Baixa as NFS-e de um CNPJ direto da API oficial ADN (mTLS).

    O fluxo antigo por navegador + extensao "Baixar NFSe" foi removido: a
    extensao mudou o modelo de download e o caminho por clique de pixel era
    fragil. Agora o download e 100% HTTP via adn_client (sem navegador).
    """

    def __init__(self, cancel_event: threading.Event | None = None) -> None:
        self.cancel_event = cancel_event
        self.base_output = Path(config.PASTA_SAIDA)
        self.base_output.mkdir(parents=True, exist_ok=True)

    def processar_cliente(
        self,
        cliente: Cliente,
        cert: CertificadoInfo,
        params: Parametros,
    ) -> tuple[int, str, int, str]:
        """Retorna (notas_emitidas, arquivo_emitidas, notas_recebidas, arquivo_recebidas)."""
        _check_cancel(self.cancel_event, f"[{cliente.cnpj}] Execucao cancelada antes do download.")
        output_dir = self.base_output / _sanitizar_nome_pasta(cliente.nome)
        output_dir.mkdir(parents=True, exist_ok=True)
        return _baixar_cliente_adn(cert, cliente, params, output_dir, self.cancel_event)

def _normalizar_cnpj(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _sanitizar_nome_pasta(nome: str) -> str:
    """Converte nome do cliente para nome de pasta válido no Windows."""
    nome = _html.unescape(nome)           # &amp; → &
    nome = re.sub(r'[<>:"/\\|?*]', '', nome)
    nome = nome.strip(". ")
    return nome[:120] or "sem_nome"


def _data_br_para_iso(data_br: str) -> str:
    """Converte DD/MM/YYYY para YYYY-MM-DD (formato de input[type='date'])."""
    partes = data_br.strip().split("/")
    if len(partes) == 3:
        return f"{partes[2]}-{partes[1]}-{partes[0]}"
    return data_br


def _mes_anterior() -> tuple[str, str]:
    hoje = date.today()
    if hoje.month == 1:
        ano_ant, mes_ant = hoje.year - 1, 12
    else:
        ano_ant, mes_ant = hoje.year, hoje.month - 1
    ultimo_dia = monthrange(ano_ant, mes_ant)[1]
    inicio = date(ano_ant, mes_ant, 1).strftime("%d/%m/%Y")
    fim = date(ano_ant, mes_ant, ultimo_dia).strftime("%d/%m/%Y")
    return inicio, fim


def _competencias_no_periodo(data_inicio_br: str, data_fim_br: str) -> set[str]:
    """Conjunto de competencias 'YYYY-MM' cobertas pelo periodo (datas BR
    'DD/MM/YYYY'). Usado para FILTRAR os DF-e baixados da API ADN por mes."""
    try:
        ini = _data_br_para_iso(data_inicio_br)  # YYYY-MM-DD
        fim = _data_br_para_iso(data_fim_br)
        ai, mi = int(ini[:4]), int(ini[5:7])
        af, mf = int(fim[:4]), int(fim[5:7])
    except Exception:
        return set()
    meses: set[str] = set()
    ano, mes = ai, mi
    while (ano, mes) <= (af, mf):
        meses.add(f"{ano:04d}-{mes:02d}")
        mes += 1
        if mes > 12:
            ano, mes = ano + 1, 1
    return meses


def _tipos_para_adn(tipos: str) -> set[str]:
    """Mapeia a opcao da GUI (emitidas/recebidas/ambas) para as pastas do ADN.
    Canceladas acompanha sempre que a nota for do tipo pedido."""
    t = (tipos or "ambas").lower()
    if t == "emitidas":
        return {"Emitidas", "Canceladas"}
    if t == "recebidas":
        return {"Recebidas", "Canceladas"}
    return {"Emitidas", "Recebidas", "Canceladas"}


def _baixar_cliente_adn(
    cert: CertificadoInfo,
    cliente: Cliente,
    params: Parametros,
    output_dir: Path,
    cancel_event: threading.Event | None = None,
) -> tuple[int, str, int, str]:
    """Baixa os DF-e do cliente direto da API ADN (mTLS) e salva em
    output_dir/<Tipo>/. Retorna (n_emitidas, arq_emitidas, n_recebidas,
    arq_recebidas) no mesmo contrato de processar_cliente."""
    import adn_client

    competencias = _competencias_no_periodo(params.data_inicio, params.data_fim) or None
    tipos = _tipos_para_adn(getattr(params, "tipos", "ambas"))
    log.info(
        "[%s] Baixando via API ADN | competencias=%s | tipos=%s",
        cliente.cnpj, sorted(competencias) if competencias else "todas", sorted(tipos),
    )

    res = adn_client.baixar_dfe(
        cert_path=cert.arquivo,
        senha=cert.senha,
        owner_cnpj=cert.documento or cliente.cnpj,
        destino=output_dir,
        competencias=competencias,
        tipos=tipos,
        cancelar=(lambda: bool(cancel_event and cancel_event.is_set())),
    )
    if res.erro:
        raise RuntimeError(f"Falha na API ADN: {res.erro}")

    log.info(
        "[%s] ADN concluido: %d emitidas, %d recebidas, %d canceladas "
        "(ultimo NSU %d; %d fora da competencia).",
        cliente.cnpj, res.emitidas, res.recebidas, res.canceladas,
        res.ultimo_nsu, res.ignorados_competencia,
    )
    arq_emit = str(output_dir / "Emitidas") if res.emitidas else ""
    arq_rec = str(output_dir / "Recebidas") if res.recebidas else ""
    return res.emitidas, arq_emit, res.recebidas, arq_rec


def preparar_parametros(
    data_inicio: str | None = None,
    data_fim: str | None = None,
    cnpjs: list[str] | None = None,
    tipos: str | None = None,
    clientes: list[dict] | None = None,
) -> Parametros:
    """Monta parametros. Se nao houver datas, usa mes anterior.

    `clientes` (opcional) e a lista da aba Clientes [{documento, nome}, ...].
    Quando informada, vira o filtro/origem de nomes da execucao no lugar do
    XLSX local — assim a automacao processa exatamente quem esta cadastrado
    na aba Clientes, e nao todos os certificados da pasta.
    """
    if not data_inicio or not data_fim:
        data_inicio, data_fim = _mes_anterior()

    # Normaliza a lista da aba Clientes em mapa documento(digitos) -> nome.
    clientes_map: dict[str, str] | None = None
    if clientes:
        clientes_map = {}
        for c in clientes:
            if not isinstance(c, dict):
                continue
            doc = _normalizar_cnpj(str(c.get("documento", "")))
            if len(doc) not in (11, 14):
                continue
            nome = str(c.get("nome", "")).strip()
            clientes_map[doc] = nome or doc
        if not clientes_map:
            clientes_map = None

    tipos_norm = str(tipos or "ambas").lower().strip()
    if tipos_norm not in ("emitidas", "recebidas", "ambas"):
        tipos_norm = "ambas"

    filtrados = None
    if cnpjs:
        vistos: set[str] = set()
        filtrados = []
        for item in cnpjs:
            cnpj = _normalizar_cnpj(item)
            if len(cnpj) == 14 and cnpj not in vistos:
                filtrados.append(cnpj)
                vistos.add(cnpj)

    def _resolver_path(valor: str) -> str:
        """Resolve caminho relativo em relacao ao diretorio de config.py."""
        if not valor:
            return valor
        p = Path(valor)
        if p.is_absolute():
            return valor
        config_dir = Path(getattr(config, "__file__", __file__)).parent
        return str(config_dir / p)

    params = Parametros(
        xlsx_path=_resolver_path(str(getattr(config, "XLSX_PATH", ""))),
        pasta_certs=str(getattr(config, "PASTA_CERTS", "")),
        pasta_saida=str(getattr(config, "PASTA_SAIDA", "")),
        data_inicio=data_inicio,
        data_fim=data_fim,
        cnpjs=filtrados,
        tipos=tipos_norm,
        clientes_map=clientes_map,
    )
    log.info(
        "Parametros: %s -> %s | CNPJs: %s | Tipos: %s | Clientes (aba): %s",
        params.data_inicio,
        params.data_fim,
        params.cnpjs or "todos",
        params.tipos,
        len(clientes_map) if clientes_map else "nao informados",
    )
    return params


def _carregar_clientes_xlsx(xlsx_path: str) -> dict[str, str]:
    """
    Carrega mapa CNPJ -> nome a partir de XLSX.
    Se a planilha nao existir ou nao puder ser lida, retorna vazio.
    """
    path = Path(xlsx_path)
    if not xlsx_path or not path.exists():
        return {}

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        log.warning("openpyxl nao disponivel para leitura de XLSX: %s", exc)
        return {}

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return {}

        header_norm = [_norm_header(cell) for cell in header]
        idx_cnpj = _find_col(
            header_norm,
            [getattr(config, "XLSX_COLUNA_CNPJ", "CNPJ"), "cnpj"],
        )
        idx_nome = _find_col(
            header_norm,
            [
                getattr(config, "XLSX_COLUNA_NOME", "NOME"),
                "nome",
                "razao",
                "empresa",
                "cliente",
            ],
        )
        if idx_cnpj is None:
            log.warning("Coluna de CNPJ nao encontrada em %s", xlsx_path)
            return {}

        clientes: dict[str, str] = {}
        for row in rows:
            if not row:
                continue
            cnpj = _normalizar_cnpj(str(row[idx_cnpj] or ""))
            if len(cnpj) != 14:
                continue
            nome = ""
            if idx_nome is not None and idx_nome < len(row):
                nome = str(row[idx_nome] or "").strip()
            clientes[cnpj] = nome or cnpj
        return clientes
    except Exception as exc:  # noqa: BLE001
        log.warning("Falha ao ler planilha %s: %s", xlsx_path, exc)
        return {}


def _norm_header(value: object) -> str:
    texto = str(value or "").strip().lower()
    return re.sub(r"\s+", "", texto)


def _find_col(headers: list[str], candidatos: list[str]) -> int | None:
    candidatos_norm = [_norm_header(c) for c in candidatos if str(c).strip()]
    for cand in candidatos_norm:
        if cand in headers:
            return headers.index(cand)
    for idx, header in enumerate(headers):
        if any(cand in header for cand in candidatos_norm):
            return idx
    return None


def _escolher_cert_mais_recente(candidatos: list) -> "CertificadoInfo":
    """Dentre certificados duplicados, retorna o com maior data de validade."""
    def _validade_key(c) -> date:
        try:
            partes = c.valido_ate.split("/")
            return date(int(partes[2]), int(partes[1]), int(partes[0]))
        except Exception:
            return date.min
    return max(candidatos, key=_validade_key)


def executar_local(
    params: Parametros,
    cancel_event: threading.Event | None = None,
) -> list[ResultadoCNPJ]:
    _check_cancel(cancel_event, "Execucao cancelada antes da leitura dos certificados.")
    certs = listar_certificados(params.pasta_certs)
    certs_unicos, duplicados = indexar_certificados_por_cnpj(certs)
    # Prioriza a lista da aba Clientes (servidor/por maquina). So cai para o
    # XLSX local quando a aba nao informou ninguem — assim a execucao respeita
    # exatamente quem esta cadastrado na aba, em vez de varrer todos os certs.
    if params.clientes_map:
        clientes_xlsx = dict(params.clientes_map)
        log.info("Usando %d cliente(s) da aba Clientes como filtro da execucao.",
                 len(clientes_xlsx))
    else:
        clientes_xlsx = _carregar_clientes_xlsx(params.xlsx_path)

    todos_cert_cnpjs = set(certs_unicos.keys()) | set(duplicados.keys())

    if params.cnpjs:
        alvo_cnpjs = list(params.cnpjs)
    elif clientes_xlsx:
        # XLSX como filtro; inclui tambem CNPJs com cert duplicado (resolvidos adiante)
        alvo_cnpjs = sorted(clientes_xlsx.keys())
        # Diagnostico: CNPJs do XLSX sem certificado correspondente
        sem_cert = sorted(set(alvo_cnpjs) - todos_cert_cnpjs)
        if sem_cert:
            log.warning(
                "ATENCAO: %d CNPJ(s) do XLSX nao tem certificado .pfx na pasta:",
                len(sem_cert),
            )
            for cnpj_faltante in sem_cert:
                # Busca CNPJs similares (primeiros 8 digitos = raiz do CNPJ)
                raiz = cnpj_faltante[:8]
                similares = [c for c in todos_cert_cnpjs if c[:8] == raiz]
                if similares:
                    log.warning(
                        "  CNPJ %s -> nao encontrado, mas existe SIMILAR: %s "
                        "(confira se o CNPJ na planilha esta correto)",
                        cnpj_faltante, ", ".join(similares),
                    )
                else:
                    log.warning(
                        "  CNPJ %s -> nao encontrado na pasta de certificados",
                        cnpj_faltante,
                    )
            # Se NENHUM CNPJ do XLSX tem cert, avisa mas respeita a planilha
            if len(sem_cert) == len(alvo_cnpjs):
                log.warning(
                    "Nenhum CNPJ do XLSX corresponde a um certificado na pasta de certs. "
                    "Verifique se PASTA_CERTS em config.py aponta para a pasta correta."
                )
    else:
        # Sem filtro: processa todos com cert valido (unicos + duplicados resolvidos)
        alvo_cnpjs = sorted(todos_cert_cnpjs)

    if not alvo_cnpjs:
        raise RuntimeError("Nenhum CNPJ encontrado para processar.")

    log.info(
        "Certificados lidos: %d | CNPJs unicos: %d | CNPJs duplicados: %d | CNPJs alvo: %d",
        len(certs),
        len(certs_unicos),
        len(duplicados),
        len(alvo_cnpjs),
    )

    resultados: list[ResultadoCNPJ] = []
    runner = NFSeRunner(cancel_event=cancel_event)
    intervalo_adn = float(getattr(config, "ADN_INTERVALO_CLIENTES_S", 1.5) or 0)

    for idx, cnpj in enumerate(alvo_cnpjs, start=1):
        _check_cancel(cancel_event, "Execucao cancelada pelo usuario.")
        # Respira entre clientes no modo ADN para nao bater no rate-limit (429).
        if idx > 1 and intervalo_adn > 0:
            time.sleep(intervalo_adn)
        nome = clientes_xlsx.get(cnpj, cnpj)
        log.info(
            "[%d/%d] Cliente XLSX: '%s' (CNPJ %s) — buscando .pfx correspondente...",
            idx, len(alvo_cnpjs), nome, cnpj,
        )

        # Resolve certificado: unico ou o mais recente entre duplicados
        cert = certs_unicos.get(cnpj)
        if cert is None and cnpj in duplicados:
            candidatos = duplicados[cnpj]
            cert = _escolher_cert_mais_recente(candidatos)
            nomes_dup = ", ".join(c.arquivo.name for c in candidatos)
            log.warning(
                "[%s] CNPJ possui %d .pfx duplicados (%s). "
                "Usando o mais recente: %s (val. %s).",
                cnpj, len(candidatos), nomes_dup,
                cert.arquivo.name, cert.valido_ate,
            )
        if cert is not None:
            log.info(
                "[%s] Pareado: '%s' -> %s (CN do certificado: %s)",
                cnpj, nome, cert.arquivo.name, cert.cn,
            )

        if cert is None:
            # Diagnostico: verifica se ha algum cert com erro para este CNPJ
            msg = (
                f"Nenhum certificado .pfx encontrado para o CNPJ {cnpj} "
                f"em {params.pasta_certs}. "
                "Verifique se o arquivo existe e se o nome segue o padrao "
                "'NOME senha SENHA.pfx'."
            )
            log.error(msg)
            resultados.append(
                ResultadoCNPJ(cnpj=cnpj, nome=nome, status="erro", erro=msg)
            )
            continue

        if not nome or nome == cnpj:
            nome = cert.nome_amigavel or cnpj

        try:
            ne, ae, nr, ar = runner.processar_cliente(Cliente(cnpj=cnpj, nome=nome), cert, params)
            resultados.append(
                ResultadoCNPJ(
                    cnpj=cnpj,
                    nome=nome,
                    status="ok",
                    notas_emitidas=ne,
                    arquivo_emitidas=ae,
                    notas_recebidas=nr,
                    arquivo_recebidas=ar,
                    thumbprint=cert.thumbprint_sha1,
                )
            )
        except Exception as exc:  # noqa: BLE001
            if isinstance(exc, ExecucaoCancelada):
                raise
            log.exception("[%s] Falha na automacao local: %s", cnpj, exc)
            resultados.append(
                ResultadoCNPJ(
                    cnpj=cnpj,
                    nome=nome,
                    status="erro",
                    thumbprint=cert.thumbprint_sha1,
                    erro=str(exc),
                )
            )

    # Fase 2: importar XMLs no Domínio Web (empresa por empresa)
    if bool(getattr(config, "DOMINIO_WEB_IMPORTAR", False)):
        _importar_dominio_web(resultados, cancel_event)

    return resultados


def _importar_dominio_web(
    resultados: list[ResultadoCNPJ],
    cancel_event: threading.Event | None,
) -> None:
    """Itera pelos resultados com notas baixadas e importa no Domínio Web."""
    try:
        from dominio_importer import DominioImporter
    except ImportError:
        log.warning("dominio_importer nao encontrado — importacao no Dominio Web ignorada.")
        return

    importer = DominioImporter()
    base_saida = Path(config.PASTA_SAIDA)

    for resultado in resultados:
        _check_cancel(cancel_event, "Execucao cancelada antes de importar no Dominio.")
        if resultado.status != "ok":
            continue
        if resultado.notas_emitidas == 0 and resultado.notas_recebidas == 0:
            log.info("[%s] Sem notas — pula importacao no Dominio Web.", resultado.cnpj)
            continue

        # Pasta nomeada com o nome do cliente (sem subpastas emitidas/recebidas)
        dir_empresa = base_saida / _sanitizar_nome_pasta(resultado.nome)
        xmls_ativos = [
            p for p in dir_empresa.rglob("*.xml")
            if "cancelad" not in str(p.parent.name).lower()
        ] if dir_empresa.exists() else []

        if not xmls_ativos:
            log.info("[%s] Nenhum XML ativo para importar.", resultado.cnpj)
            continue

        try:
            log.info("[%s] Iniciando importacao no Dominio Web (%d XMLs)...", resultado.cnpj, len(xmls_ativos))
            ok = importer.importar(resultado.cnpj, resultado.nome, [dir_empresa])
            resultado.importado_dominio = ok
            if ok:
                log.info("[%s] Importacao no Dominio Web concluida.", resultado.cnpj)
            else:
                log.warning("[%s] Importacao no Dominio Web falhou ou foi ignorada.", resultado.cnpj)
        except Exception as exc:  # noqa: BLE001
            log.error("[%s] Erro ao importar no Dominio Web: %s", resultado.cnpj, exc)


def montar_mensagem(resultados: list[ResultadoCNPJ], params: Parametros) -> dict:
    """Formata o resumo final para envio por e-mail."""
    ok = sum(1 for r in resultados if r.status == "ok")
    erro = sum(1 for r in resultados if r.status == "erro")
    total = len(resultados)
    total_emitidas = sum(r.notas_emitidas for r in resultados if r.status == "ok")
    total_recebidas = sum(r.notas_recebidas for r in resultados if r.status == "ok")
    total_notas = total_emitidas + total_recebidas

    linhas: list[str] = []
    for r in resultados:
        if r.status == "ok":
            detalhe = f"emitidas={r.notas_emitidas} | recebidas={r.notas_recebidas}"
            if r.arquivo_emitidas:
                detalhe += f" | arq_emit={Path(r.arquivo_emitidas).name}"
            if r.arquivo_recebidas:
                detalhe += f" | arq_rec={Path(r.arquivo_recebidas).name}"
            dominio = " | dominio=OK" if r.importado_dominio else ""
            linhas.append(f"- OK   | {r.cnpj} | {r.nome} | {detalhe}{dominio}")
        else:
            linhas.append(f"- ERRO | {r.cnpj} | {r.nome}")
            linhas.append(f"  Erro: {r.erro}")

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    assunto = f"NFSe Automacao - Relatorio ({ok} OK, {erro} erro)"

    mensagem = "\n".join(
        [
            "NFSe Automacao - Relatorio",
            f"Horario: {agora}",
            f"Periodo: {params.data_inicio} -> {params.data_fim}",
            "",
            f"Total CNPJs: {total}",
            f"Sucesso: {ok}",
            f"Erros: {erro}",
            f"Notas emitidas: {total_emitidas}",
            f"Notas recebidas: {total_recebidas}",
            f"Total notas: {total_notas}",
            "",
            "Detalhes:",
            *(linhas or ["- Nenhum item processado."]),
        ]
    )

    return {
        "assunto": assunto,
        "mensagem": mensagem,
        "total": total,
        "ok": ok,
        "erro": erro,
        "houve_erro": erro > 0,
    }


def montar_mensagem_erro(erro: str) -> dict:
    """Formata a mensagem de falha critica."""
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    return {
        "assunto": "NFSe Automacao - Falha critica",
        "mensagem": (
            "NFSe Automacao - Falha critica\n\n"
            f"Erro: {erro}\n"
            f"Horario: {agora}"
        ),
    }


def _destinatarios_email() -> list[str]:
    return [
        item.strip()
        for item in str(getattr(config, "ZOHO_EMAIL_TO", "")).split(",")
        if item.strip()
    ]


def notificar_email(assunto: str, mensagem: str) -> None:
    """Envia e-mail via SMTP do Zoho Mail. E OPCIONAL: se o SMTP/destinatarios
    nao estiverem configurados, apenas registra e segue (nao derruba a execucao)."""
    destinatarios = _destinatarios_email()
    if not (str(getattr(config, "ZOHO_SMTP_HOST", "")).strip()
            and str(getattr(config, "ZOHO_EMAIL_FROM", "")).strip()
            and destinatarios):
        log.info("E-mail nao configurado (ZOHO_SMTP_*/ZOHO_EMAIL_*) — relatorio nao enviado.")
        return

    email = EmailMessage()
    email["Subject"] = assunto
    email["From"] = config.ZOHO_EMAIL_FROM
    email["To"] = ", ".join(destinatarios)
    email.set_content(mensagem)

    try:
        with smtplib.SMTP(
            config.ZOHO_SMTP_HOST,
            config.ZOHO_SMTP_PORT,
            timeout=30,
        ) as smtp:
            smtp.ehlo()
            if config.ZOHO_SMTP_USE_TLS:
                smtp.starttls()
                smtp.ehlo()
            smtp.login(config.ZOHO_SMTP_USER, config.ZOHO_SMTP_PASSWORD)
            smtp.send_message(email)
        log.info("E-mail enviado com sucesso.")
    except Exception as exc:  # noqa: BLE001
        log.error("Falha ao enviar e-mail: %s", exc)


def executar(
    data_inicio: str | None = None,
    data_fim: str | None = None,
    cnpjs: list[str] | None = None,
    cancel_event: threading.Event | None = None,
) -> None:
    """Executa o fluxo completo da automacao local."""
    log.info("=" * 60)
    log.info("Iniciando automacao NFSe (API oficial ADN)")

    try:
        params = preparar_parametros(data_inicio, data_fim, cnpjs)
        resultados = executar_local(params, cancel_event=cancel_event)

        resumo = montar_mensagem(resultados, params)
        log.info(
            "Resultado: %d ok / %d erro / %d total",
            resumo["ok"],
            resumo["erro"],
            resumo["total"],
        )
        notificar_email(resumo["assunto"], resumo["mensagem"])
    except ExecucaoCancelada as exc:
        log.warning("Automacao cancelada: %s", exc)
        raise
    except Exception as exc:  # noqa: BLE001
        log.error("Falha critica: %s", exc, exc_info=True)
        msg = montar_mensagem_erro(str(exc))
        notificar_email(msg["assunto"], msg["mensagem"])

    log.info("Automacao NFSe finalizada")
    log.info("=" * 60)


if __name__ == "__main__":
    executar()
