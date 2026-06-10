"""
FastAPI backend para NFSe Automacao.
Expõe endpoints REST para o frontend Electron.
"""
import hmac
import importlib
import logging
import os
import re
import sys
import threading
import uuid
from collections import deque
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import requests
import uvicorn

# =============================================================================
# Logging — PRIMEIRO, antes de qualquer import que possa usar logger
# =============================================================================
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Suporte a bundle PyInstaller: config.py e módulos ficam ao lado do server.exe
if getattr(sys, "frozen", False):
    _app_dir = str(Path(sys.executable).parent)
    if _app_dir not in sys.path:
        sys.path.insert(0, _app_dir)
else:
    sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from config_defaults import apply_defaults, default_values, ensure_config_file
except ImportError as e:
    logger.warning(f"Falha ao importar config_defaults: {e}")
    apply_defaults = None
    default_values = lambda: {}
    ensure_config_file = lambda: Path(__file__).parent.parent / "config.py"

try:
    ensure_config_file()
except Exception as e:
    logger.warning("Falha ao criar config.py padrao: %s", e)

try:
    import config
    if apply_defaults:
        apply_defaults(config)
except ImportError:
    logger.warning("Falha ao importar config")
    class config:
        pass
    if apply_defaults:
        apply_defaults(config, create_dirs=False)

try:
    from cert_reader import listar_certificados, indexar_certificados_por_cnpj
except ImportError as e:
    logger.warning(f"Falha ao importar cert_reader: {e}. Usando fallback.")
    def listar_certificados(*args, **kwargs):
        return []
    def indexar_certificados_por_cnpj(*args, **kwargs):
        return {}, {}

try:
    from nfse_automacao import ExecucaoCancelada, executar_local, preparar_parametros
except ImportError as e:
    logger.warning(f"Falha ao importar nfse_automacao: {e}. Usando fallback.")
    ExecucaoCancelada = RuntimeError
    def executar_local(*args, **kwargs):
        return []
    def preparar_parametros(*args, **kwargs):
        return {}

import importlib.util as _ilu
import time as _time

def _load_license_module():
    # Procura license.py em varios locais: ao lado de server.py (dev) e, no app
    # empacotado, ao lado do server.exe / dentro do bundle (_MEIPASS).
    _cands = [Path(__file__).parent / "license.py"]
    if getattr(sys, "frozen", False):
        _cands.append(Path(sys.executable).parent / "license.py")
        _meipass = getattr(sys, "_MEIPASS", None)
        if _meipass:
            _cands.append(Path(_meipass) / "license.py")
    _p = next((c for c in _cands if c.exists()), None)
    if _p is None:
        logger.error("license.py nao encontrado. Procurei em: %s",
                     ", ".join(str(c) for c in _cands))
        return None
    try:
        _spec = _ilu.spec_from_file_location("license", _p)
        _mod  = _ilu.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        return _mod
    except Exception as e:
        logger.error("Falha ao carregar license.py (%s): %s", _p, e)
        return None

_lic = _load_license_module()
if _lic:
    license_activate   = _lic.activate
    check_license      = _lic.check_license
    validate_license_key = _lic.validate_key
    load_key           = _lic.load_key
    license_deactivate = _lic.deactivate
    logger.info("Módulo de licença carregado.")
else:
    logger.error("LICENÇA: módulo não encontrado — execução BLOQUEADA.")
    def license_activate(key): return False, "Módulo de licença não encontrado."
    def check_license(key=None): return False, "Módulo de licença não encontrado."
    def validate_license_key(key): return False, "Módulo de licença não encontrado."
    def load_key(): return None
    def license_deactivate(): pass

# =============================================================================
# Job management
# =============================================================================
@dataclass
class Job:
    id: str
    status: str  # "running", "ok", "cancelado", "erro"
    logs: deque = None
    thread: Optional[threading.Thread] = None
    cancel_event: Optional[threading.Event] = None
    resultado: Optional[dict] = None
    created_at: float = 0.0

    def __post_init__(self):
        if self.logs is None:
            self.logs = deque(maxlen=2000)
        if self.cancel_event is None:
            self.cancel_event = threading.Event()
        if self.created_at == 0.0:
            self.created_at = _time.time()

JOBS = {}
_MAX_JOBS = 50
_JOB_TTL_S = 7200  # 2 horas


def _prune_jobs() -> None:
    """Remove jobs finalizados mais antigos quando o limite é atingido."""
    if len(JOBS) <= _MAX_JOBS:
        return
    cutoff = _time.time() - _JOB_TTL_S
    to_remove = [
        jid for jid, j in JOBS.items()
        if j.status != "running" and j.created_at < cutoff
    ]
    for jid in to_remove:
        del JOBS[jid]
    # Se ainda acima do limite, remove os mais antigos independente do TTL
    if len(JOBS) > _MAX_JOBS:
        finished = sorted(
            [(jid, j) for jid, j in JOBS.items() if j.status != "running"],
            key=lambda x: x[1].created_at,
        )
        for jid, _ in finished[:len(JOBS) - _MAX_JOBS]:
            del JOBS[jid]

class JobLogHandler(logging.Handler):
    """Handler que adiciona logs a um Job específico."""
    def __init__(self, job_id: str):
        super().__init__()
        self.job_id = job_id
        self.setFormatter(
            logging.Formatter(
                fmt="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
                datefmt="%H:%M:%S",
            )
        )

    def emit(self, record: logging.LogRecord) -> None:
        if self.job_id in JOBS:
            msg = self.format(record)
            JOBS[self.job_id].logs.append({
                "timestamp": record.created,
                "level": record.levelname,
                "message": msg,
            })

# =============================================================================
# FastAPI app
# =============================================================================
app = FastAPI(title="NFSe Automacao API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:5173", "http://localhost:5173", "app://"],
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "x-api-token", "x-license-key"],
)

# Token gerado pelo Electron e passado via env var NFSE_API_TOKEN.
# Se não definido (modo dev sem Electron), a auth é desativada.
_API_TOKEN = os.environ.get("NFSE_API_TOKEN", "")

def _is_license_exempt_path(method: str, path: str) -> bool:
    if path == "/health":
        return True
    if method == "GET" and re.fullmatch(r"/executar/[^/]+/status", path):
        return True
    if method == "POST" and re.fullmatch(r"/executar/[^/]+/cancelar", path):
        return True
    return False

@app.middleware("http")
async def _require_token(request: Request, call_next):
    if request.url.path == "/health":
        return await call_next(request)
    if _API_TOKEN:
        token = request.headers.get("x-api-token", "")
        if not token or not hmac.compare_digest(token.encode(), _API_TOKEN.encode()):
            return JSONResponse(status_code=401, content={"error": "Não autorizado"})

    if request.method != "OPTIONS" and not _is_license_exempt_path(request.method, request.url.path):
        license_key = request.headers.get("x-license-key", "")
        licensed, message = validate_license_key(license_key)
        if not licensed:
            return JSONResponse(
                status_code=402,
                content={"licensed": False, "detail": message, "message": message},
            )
    return await call_next(request)

# =============================================================================
# Endpoints
# =============================================================================

@app.get("/health")
async def health():
    """Health check — Electron aguarda isso para saber que o backend está pronto."""
    return {"ok": True}

@app.get("/config")
async def get_config():
    """Retorna seções + valores atuais de config.py."""
    sections = [
        {
            "title": "Caminhos locais",
            "fields": [
                ("XLSX_PATH", "Planilha de clientes", "path"),
                ("PASTA_CERTS", "Pasta de certificados", "dir"),
                ("PASTA_SAIDA", "Pasta de saida", "dir"),
                ("CHROME_USER_DATA_DIR", "Perfil Chrome", "dir"),
                ("CHROME_EXTENSION_DIR", "Pasta da extensao (opcional)", "dir"),
                ("CHROME_EXTENSION_ID", "ID da extensao NFSe (auto-detectado se vazio)", "text"),
            ],
        },
        {
            "title": "Portal NFSe / Playwright",
            "fields": [
                ("NFSE_LOGIN_URL", "URL de login", "text"),
                ("NFSE_EMITIDAS_URL", "URL de notas emitidas", "text"),
                ("NFSE_RECEBIDAS_URL", "URL de notas recebidas", "text"),
                ("AUTOSELECT_CERTIFICATE_PATTERNS", "AutoSelectCertificateForUrls", "text"),
                ("CHROME_CHANNEL", "Canal do browser", "text"),
                ("CHROME_EXECUTABLE_PATH", "Executavel Chrome (opcional)", "path"),
                ("PLAYWRIGHT_HEADLESS", "Headless (True/False)", "text"),
                ("PLAYWRIGHT_TIMEOUT_MS", "Timeout padrao (ms)", "int"),
                ("PLAYWRIGHT_LOGIN_TIMEOUT_S", "Timeout login (s)", "int"),
                ("PLAYWRIGHT_DOWNLOAD_TIMEOUT_S", "Timeout download (s)", "int"),
            ],
        },
        {
            "title": "Seletores e extensao",
            "fields": [
                ("NFSE_SELECTOR_LOGIN_OK", "Seletor de login OK", "text"),
                ("NFSE_SELECTOR_BOTAO_CERTIFICADO", "Botao acesso certificado", "text"),
                ("NFSE_SELECTOR_DATA_INICIO", "Campo data inicio", "text"),
                ("NFSE_SELECTOR_DATA_FIM", "Campo data fim", "text"),
                ("NFSE_SELECTOR_BOTAO_FILTRAR", "Botao filtrar", "text"),
                ("NFSE_SELECTOR_LINHAS_NOTAS", "Linhas da tabela de notas", "text"),
                ("NFSE_SELECTOR_TEXTO_SEM_NOTAS", "Textos de sem notas", "text"),
                ("NFSE_SELECTOR_BOTAO_BAIXAR", "Botao da extensao", "text"),
                ("NFSE_ATALHO_EXTENSAO", "Atalho da extensao", "text"),
            ],
        },
        {
            "title": "Licencas (admin — so na maquina do administrador)",
            "fields": [
                ("LICENSE_ADMIN_URL", "URL admin do servidor de licencas", "text"),
                ("LICENSE_ADMIN_TOKEN", "Token admin (para ver as maquinas no Modo Dev)", "text"),
            ],
        },
        {
            "title": "E-mail (Zoho SMTP)",
            "fields": [
                ("ZOHO_SMTP_HOST", "Host SMTP", "text"),
                ("ZOHO_SMTP_PORT", "Porta SMTP", "int"),
                ("ZOHO_SMTP_USER", "Usuario", "text"),
                ("ZOHO_SMTP_PASSWORD", "Senha", "secret"),
                ("ZOHO_EMAIL_FROM", "Remetente", "text"),
                ("ZOHO_EMAIL_TO", "Destinatario(s)", "text"),
            ],
        },
    ]

    values = {}
    for section in sections:
        for key, label, tipo in section["fields"]:
            val = getattr(config, key, "")
            values[key] = "••••••••" if tipo == "secret" else str(val)

    return {"sections": sections, "values": values}

@app.post("/config")
async def set_config(body: dict):
    """Salva valores em config.py (com backup .bak)."""
    config_path = (
        Path(sys.executable).parent if getattr(sys, "frozen", False)
        else Path(__file__).parent.parent
    ) / "config.py"
    if not config_path.exists():
        try:
            config_path = ensure_config_file(config_path.parent)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Falha ao criar config.py: {e}")
    if not config_path.exists():
        raise HTTPException(status_code=404, detail="config.py não encontrado")

    try:
        original = config_path.read_text(encoding="utf-8")
        novo = original
        defaults = default_values()

        for chave, valor in body.items():
            if not hasattr(config, chave):
                continue
            if str(valor).strip() == "" and defaults.get(chave) not in ("", None):
                valor = defaults[chave]
            atual = getattr(config, chave)
            if isinstance(atual, bool):
                literal = "True" if str(valor).lower() in ("true", "1", "sim") else "False"
            elif isinstance(atual, int):
                literal = str(int(valor))
            else:
                escaped = (str(valor)
                          .replace("\\", "\\\\")
                          .replace('"', '\\"')
                          .replace("\n", "\\n")
                          .replace("\r", "\\r"))
                literal = f'r"{valor}"' if ("\\" in str(valor) and '"' not in str(valor) and "\n" not in str(valor) and "\r" not in str(valor)) else f'"{escaped}"'

            pattern = re.compile(
                rf"^(?P<prefix>{re.escape(chave)}\s*=\s*)(?P<value>.+?)(?P<suffix>\s*(?:#.*)?)$",
                re.MULTILINE,
            )
            novo, count = pattern.subn(
                lambda m: f"{m.group('prefix')}{literal}{m.group('suffix')}",
                novo,
                count=1,
            )
            if count == 0:
                novo = f"{novo.rstrip()}\n{chave} = {literal}\n"

        backup = config_path.with_name(f"config.py.bak.{int(_time.time())}")
        backup.write_text(original, encoding="utf-8")

        # Remove backups mais antigos (mantém últimos 5)
        old_backups = sorted(config_path.parent.glob("config.py.bak.*"),
                             key=lambda p: p.stat().st_mtime)
        for old in old_backups[:-5]:
            try:
                old.unlink()
            except Exception:
                pass
        config_path.write_text(novo, encoding="utf-8")

        importlib.reload(config)
        if apply_defaults:
            apply_defaults(config)

        return {"ok": True, "backupPath": str(backup)}
    except Exception as e:
        logger.exception("Falha ao salvar config")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/paths/status")
async def paths_status():
    """Diagnostico dos caminhos + sugestoes auto-detectadas (1a execucao).

    Usado pelo frontend para decidir se mostra o onboarding e ja pre-preencher
    as pastas detectadas (certificados no Google Drive, DOMINIO WEB local).
    """
    try:
        import path_finder as pf
    except Exception:
        pf = None

    certs_path = str(getattr(config, "PASTA_CERTS", "") or "")
    saida_path = str(getattr(config, "PASTA_SAIDA", "") or "")

    def _certs_info(p: str) -> dict:
        pp = Path(p) if p else None
        exists = bool(pp and pp.is_dir())
        pfx = 0
        if exists:
            try:
                pfx = sum(
                    1 for f in pp.iterdir()
                    if f.is_file() and f.suffix.lower() == ".pfx"
                )
            except OSError:
                pfx = 0
        return {"path": p, "exists": exists, "pfxCount": pfx, "ok": exists and pfx > 0}

    def _dir_info(p: str) -> dict:
        pp = Path(p) if p else None
        exists = bool(pp and pp.is_dir())
        return {"path": p, "exists": exists, "ok": exists}

    certs = _certs_info(certs_path)
    saida = _dir_info(saida_path)

    sug_certs = sug_saida = ""
    if pf:
        try:
            achado_certs = pf.find_certs_folder()
            sug_certs = str(achado_certs) if achado_certs else ""
            achado_dom = pf.find_dominio_web_folder()
            if achado_dom:
                sug_saida = str(Path(achado_dom) / pf.SIMPLES_NACIONAL_NAME)
        except Exception:
            pass

    return {
        "certs": certs,
        "saida": saida,
        "needsSetup": not certs["ok"],
        "suggestions": {
            "PASTA_CERTS": sug_certs or (certs_path if certs["ok"] else ""),
            "PASTA_SAIDA": sug_saida or saida_path,
        },
    }


@app.get("/admin/maquinas")
async def admin_maquinas():
    """Status do servidor de licencas + lista de licencas/maquinas ativas.

    - 'online': o servidor respondeu (conectividade).
    - 'configurado': LICENSE_ADMIN_TOKEN setado nesta maquina (mostra os dados).
    - 'itens': licencas com suas maquinas; 'maquinas': lista achatada (1 por
      maquina ativa) para exibir direto no Modo Dev.
    """
    url = (
        os.environ.get("NFSE_LICENSE_ADMIN_URL", "").strip()
        or str(getattr(config, "LICENSE_ADMIN_URL", "") or "").strip()
    )
    token = str(getattr(config, "LICENSE_ADMIN_TOKEN", "") or "").strip()
    base = url.split("/api/")[0] if "/api/" in url else url
    out = {"serverUrl": base, "online": False, "configurado": bool(token),
           "ok": False, "total": 0, "itens": [], "maquinas": [], "erro": ""}
    if not url:
        out["erro"] = "LICENSE_ADMIN_URL nao configurada."
        return out
    try:
        import requests
        if token:
            r = requests.post(
                url, json={"op": "list"},
                headers={"x-admin-token": token, "Content-Type": "application/json"},
                timeout=15,
            )
            out["online"] = True
            data = r.json()
            out["ok"] = bool(data.get("ok"))
            out["total"] = data.get("total", 0)
            out["itens"] = data.get("items", [])
            if not out["ok"]:
                out["erro"] = str(data.get("message", ""))
            # Achata as maquinas (1 entrada por maquina com chave ativa)
            for it in out["itens"]:
                for mid, m in (it.get("machines") or {}).items():
                    machine_status = (m or {}).get("status") or it.get("status", "active")
                    if it.get("status") == "blocked":
                        machine_status = "blocked"
                    out["maquinas"].append({
                        "cliente": it.get("clientName") or it.get("keyFmt"),
                        "machineId": mid,
                        "status": machine_status,
                        "expiresAt": it.get("expiresAt"),
                        "lastSeen": (m or {}).get("lastSeen"),
                        "firstSeen": (m or {}).get("firstSeen"),
                    })
            out["maquinas"].sort(key=lambda x: x.get("lastSeen") or "", reverse=True)
        else:
            # Sem token: so verifica se o servidor esta no ar.
            requests.get(base, timeout=10)
            out["online"] = True
        return out
    except Exception as e:
        out["online"] = False
        out["erro"] = str(e)
        return out


@app.get("/certificados")
async def count_certificados(incluir_lista: bool = False):
    """Lista certificados da PASTA_CERTS."""
    try:
        pasta_raw = str(getattr(config, "PASTA_CERTS", "")).strip()
        if not pasta_raw:
            raise ValueError("PASTA_CERTS não configurada")

        pasta = Path(pasta_raw)
        if not pasta.exists():
            raise FileNotFoundError(f"Pasta não encontrada: {pasta}")

        certs = listar_certificados(pasta)
        mapa_unico, duplicados = indexar_certificados_por_cnpj(certs)

        validos = [c for c in certs if not c.erro]
        ok = len(validos)
        erro = len(certs) - ok
        ecnpj = sum(1 for c in validos if len(c.documento) == 14)
        ecpf = sum(1 for c in validos if len(c.documento) == 11)
        sem_doc = sum(1 for c in validos if len(c.documento) not in (11, 14))
        dup_arquivos = sum(len(itens) for itens in duplicados.values())

        resultado = {
            "total": len(certs),
            "validos": ok,
            "erros": erro,
            "ecnpj": ecnpj,
            "ecpf": ecpf,
            "semDoc": sem_doc,
            "cnpjsUnicos": len(mapa_unico),
            "cnpjsDuplicados": len(duplicados),
            "arquivosDuplicados": dup_arquivos,
            "path": str(pasta),
        }

        if incluir_lista:
            resultado["lista"] = [
                {
                    "arquivo": c.arquivo.name,
                    "nomeAmigavel": c.nome_amigavel,
                    "cn": c.cn,
                    "documento": c.documento,
                    "validoAte": c.valido_ate,
                    "erro": c.erro or "",
                }
                for c in certs
            ]

        return resultado
    except Exception as e:
        logger.exception("Falha ao contar certificados")
        raise HTTPException(status_code=500, detail=str(e))

def _resolver_xlsx_path() -> Path:
    """Resolve XLSX_PATH relativo ao diretorio do config.py."""
    raw = str(getattr(config, "XLSX_PATH", "clientes.xlsx")).strip() or "clientes.xlsx"
    p = Path(raw)
    if p.is_absolute():
        return p
    config_dir = Path(getattr(config, "__file__", __file__)).parent
    return config_dir / raw


def _normalizar_clientes(clientes) -> list[dict]:
    out = []
    for c in clientes or []:
        if not isinstance(c, dict):
            continue
        doc = re.sub(r"\D", "", str(c.get("documento", "")).strip())[:32]
        nome = str(c.get("nome", "")).strip()[:200]
        if doc or nome:
            out.append({"documento": doc, "nome": nome})
    return out


def _ler_clientes_local() -> dict:
    from openpyxl import load_workbook

    xlsx_path = _resolver_xlsx_path()
    if not xlsx_path.exists():
        return {"clientes": [], "path": str(xlsx_path), "source": "local"}

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return {"clientes": [], "path": str(xlsx_path), "source": "local"}

    header = [str(c or "").strip().lower() for c in rows[0]]
    idx_doc = next((i for i, h in enumerate(header)
                    if "cnpj" in h or "cpf" in h or "doc" in h), 0)
    idx_nome = next((i for i, h in enumerate(header)
                     if "nome" in h or "razao" in h or "cliente" in h), 1)

    clientes = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        doc = str(row[idx_doc] or "").strip() if idx_doc < len(row) else ""
        nome = str(row[idx_nome] or "").strip() if idx_nome < len(row) else ""
        if doc or nome:
            clientes.append({"documento": doc, "nome": nome})

    return {"clientes": _normalizar_clientes(clientes), "path": str(xlsx_path), "source": "local"}


def _salvar_clientes_local(clientes: list[dict]) -> dict:
    from openpyxl import Workbook

    xlsx_path = _resolver_xlsx_path()
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["CNPJ", "NOME"])
    for c in clientes:
        ws.append([c.get("documento", ""), c.get("nome", "")])
    wb.save(xlsx_path)
    wb.close()

    return {"ok": True, "salvos": len(clientes), "path": str(xlsx_path), "source": "local"}


def _clientes_remotos_ativos() -> bool:
    value = os.environ.get("NFSE_CLIENTES_REMOTE", "1").strip().lower()
    return value not in {"0", "false", "no", "off"}


def _license_server_url() -> str:
    url = (
        os.environ.get("NFSE_LICENSE_SERVER_URL", "").strip()
        or str(getattr(_lic, "LICENSE_SERVER_URL", "") or "").strip()
        or "https://license-server-sigma-topaz.vercel.app"
    )
    return url.rstrip("/")


def _clientes_remote_url() -> str:
    return (
        os.environ.get("NFSE_CLIENTES_URL", "").strip()
        or f"{_license_server_url()}/api/clientes"
    )


def _clientes_machine_id() -> str:
    if _lic and hasattr(_lic, "get_machine_id"):
        return _lic.get_machine_id()
    raise RuntimeError("Identificador desta maquina indisponivel.")


def _clientes_client_secret() -> str:
    return (
        os.environ.get("NFSE_CLIENT_SECRET", "").strip()
        or str(getattr(_lic, "_CLIENT_SECRET", "") or "").strip()
    )


def _clientes_remote_call(op: str, license_key: str, clientes: list[dict] | None = None) -> dict:
    if not license_key:
        raise RuntimeError("Licenca nao informada.")

    payload = {
        "op": op,
        "key": license_key,
        "machine_id": _clientes_machine_id(),
    }
    if clientes is not None:
        payload["clientes"] = clientes

    headers = {"Content-Type": "application/json"}
    client_secret = _clientes_client_secret()
    if client_secret:
        headers["x-client-secret"] = client_secret

    resp = requests.post(
        _clientes_remote_url(),
        json=payload,
        headers=headers,
        timeout=10,
    )
    try:
        data = resp.json()
    except Exception:
        data = {}
    if not resp.ok or data.get("ok") is False:
        msg = data.get("message") or data.get("detail") or resp.text or f"HTTP {resp.status_code}"
        raise RuntimeError(str(msg))
    return data


@app.get("/clientes")
async def listar_clientes(request: Request):
    """Le clientes por maquina no servidor; usa XLSX local como fallback/migracao."""
    try:
        if _clientes_remotos_ativos():
            try:
                data = _clientes_remote_call("get", request.headers.get("x-license-key", ""))
                clientes = _normalizar_clientes(data.get("clientes") or [])
                if clientes or data.get("updatedAt"):
                    return {
                        "clientes": clientes,
                        "path": data.get("path") or "",
                        "location": data.get("location") or "Servidor de licencas",
                        "source": "server",
                        "machineId": data.get("machine_id") or _clientes_machine_id(),
                        "updatedAt": data.get("updatedAt"),
                    }

                local = _ler_clientes_local()
                if local.get("clientes"):
                    local["location"] = (
                        f"Arquivo local: {local.get('path')} "
                        "(clique em Salvar para enviar ao servidor desta maquina)"
                    )
                    local["sync_hint"] = "Clique em Salvar para enviar estes clientes ao servidor desta maquina."
                    return local

                return {
                    "clientes": [],
                    "path": data.get("path") or "",
                    "location": data.get("location") or "Servidor de licencas",
                    "source": "server",
                    "machineId": data.get("machine_id") or _clientes_machine_id(),
                    "updatedAt": data.get("updatedAt"),
                }
            except Exception as remote_err:
                logger.warning("Falha ao carregar clientes do servidor: %s", remote_err)

        local = _ler_clientes_local()
        if not local.get("location"):
            local["location"] = f"Arquivo local: {local.get('path')}"
        return local
    except Exception as e:
        logger.exception("Falha ao listar clientes")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/clientes")
async def salvar_clientes(request: Request, payload: dict):
    """Grava clientes por maquina no servidor; local so se NFSE_CLIENTES_REMOTE=0."""
    try:
        clientes = _normalizar_clientes(payload.get("clientes") or [])

        if _clientes_remotos_ativos():
            try:
                data = _clientes_remote_call(
                    "save",
                    request.headers.get("x-license-key", ""),
                    clientes,
                )
                return {
                    "ok": True,
                    "salvos": int(data.get("salvos", len(clientes))),
                    "path": data.get("path") or "",
                    "location": data.get("location") or "Servidor de licencas",
                    "source": "server",
                    "machineId": data.get("machine_id") or _clientes_machine_id(),
                    "updatedAt": data.get("updatedAt"),
                }
            except Exception as remote_err:
                logger.exception("Falha ao salvar clientes no servidor")
                raise HTTPException(
                    status_code=502,
                    detail=f"Falha ao salvar clientes no servidor: {remote_err}",
                )

        return _salvar_clientes_local(clientes)
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        logger.exception("Falha ao salvar clientes")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/license/status")
async def license_status():
    """Endpoint legado: licença agora é validada no Electron contra a Vercel."""
    return {
        "licensed": False,
        "message": "Licenca local desativada. O aplicativo valida diretamente no servidor da Vercel.",
        "key_hint": None,
        "remoteOnly": True,
    }


@app.post("/license/activate")
async def license_activate_endpoint(body: dict):
    """Endpoint legado: ativação local foi desativada."""
    raise HTTPException(
        status_code=410,
        detail="Ativacao local desativada. Use a validacao direta do aplicativo com a Vercel.",
    )


@app.post("/license/deactivate")
async def license_deactivate_endpoint():
    """Remove a chave de licença local."""
    license_deactivate()
    return {"ok": True}


@app.post("/executar")
async def start_execution(body: dict):
    """Inicia execução assíncrona. Retorna { jobId }."""
    _prune_jobs()

    try:
        data_inicio = body.get("dataInicio")
        data_fim = body.get("dataFim")
        cnpjs = body.get("cnpjs")
        tipos = body.get("tipos")  # "emitidas" | "recebidas" | "ambas"

        job_id = str(uuid.uuid4())[:8]
        job = Job(id=job_id, status="running")
        JOBS[job_id] = job

        handler = JobLogHandler(job_id)
        root_logger = logging.getLogger()
        root_logger.addHandler(handler)

        def run_job():
            try:
                params = preparar_parametros(data_inicio, data_fim, cnpjs, tipos)
                resultado = executar_local(params, cancel_event=job.cancel_event)
                job.status = "ok"
                job.resultado = {
                    "resultados": [asdict(r) for r in resultado],
                    "total": len(resultado),
                    "ok": sum(1 for r in resultado if r.status == "ok"),
                    "erro": sum(1 for r in resultado if r.status == "erro"),
                }
            except ExecucaoCancelada:
                job.status = "cancelado"
            except Exception as e:
                logger.exception("Erro na execução")
                job.status = "erro"
                job.resultado = {"erro": str(e)}
            finally:
                root_logger.removeHandler(handler)

        job.thread = threading.Thread(target=run_job, daemon=True)
        job.thread.start()

        return {"jobId": job_id}
    except Exception as e:
        logger.exception("Falha ao iniciar execução")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/executar/{job_id}/status")
async def get_job_status(job_id: str):
    """Retorna status de um job."""
    if job_id not in JOBS:
        raise HTTPException(status_code=404, detail="Job não encontrado")

    job = JOBS[job_id]
    logs = list(job.logs)

    resultado = {
        "jobId": job_id,
        "status": job.status,
        "logs": logs,
    }

    if job.resultado:
        resultado["resultado"] = job.resultado

    return resultado

@app.post("/executar/{job_id}/cancelar")
async def cancel_job(job_id: str):
    """Cancela um job em execução."""
    if job_id not in JOBS:
        raise HTTPException(status_code=404, detail="Job não encontrado")

    job = JOBS[job_id]
    if job.status == "running":
        job.cancel_event.set()
        return {"ok": True, "message": "Cancelamento solicitado"}
    return {"ok": False, "message": "Job não está em execução"}

# =============================================================================
# Entry point
# =============================================================================
if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=17432, log_level="info")
